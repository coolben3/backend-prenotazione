from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook
import csv
import os
import shutil

app = Flask(__name__)
CORS(app)

EXCEL_PATH = "template.xlsx"
BACKUP_PATH = "backup_template.xlsx"
CSV_PATH = "storico.csv"

# Campi del form → celle Excel
CELL_MAP = {
    "nome_cognome": "B6",
    "tratta": "E13",
    "giorno": "E16",
    "mese": "E17",
    "adulti": "M6",
    "bambini": "M7",
    "neonati": "M8",
    "moto": "M9",
    "auto": "M10",
    "orario": "D15",
    "email": "B22"
}

@app.route("/invia", methods=["POST"])
def ricevi_dati():
    data = request.get_json()

    if not data:
        return jsonify({"error": "Nessun dato ricevuto"}), 400

    try:
        # 🔁 Backup automatico
        shutil.copy(EXCEL_PATH, BACKUP_PATH)

        # 📥 Scrive nel template originale (con formule attive)
        wb = load_workbook(EXCEL_PATH, keep_vba=True)  # usa keep_vba=True se hai macro
        ws = wb.active

        for campo, cella in CELL_MAP.items():
            valore = data.get(campo)
            if valore is not None:
                ws[cella] = valore

        wb.save(EXCEL_PATH)

        # 📝 Salva anche lo storico in CSV
        aggiungi_a_csv(data)

        return jsonify({"message": "Dati salvati con successo ✅"}), 200

    except Exception as e:
        return jsonify({"error": "Errore durante la scrittura", "details": str(e)}), 500

def aggiungi_a_csv(data):
    headers = list(data.keys())
    file_esiste = os.path.isfile(CSV_PATH)

    with open(CSV_PATH, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        if not file_esiste:
            writer.writeheader()
        writer.writerow(data)

@app.route("/download", methods=["GET"])
def scarica_excel():
    try:
        return send_file(
            EXCEL_PATH,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='biglietto_compilato.xlsx'
        )
    except Exception as e:
        return jsonify({"error": "Errore durante il download", "details": str(e)}), 500

@app.route("/storico", methods=["GET"])
def scarica_csv():
    try:
        return send_file(
            CSV_PATH,
            mimetype='text/csv',
            as_attachment=True,
            download_name='storico.csv'
        )
    except Exception as e:
        return jsonify({"error": "Errore durante il download del CSV", "details": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=3000)
